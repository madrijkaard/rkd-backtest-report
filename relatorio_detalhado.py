import os
import glob
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import load_workbook
from tqdm import tqdm

MONTHS_EN = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December"
}

def generate_detailed_reports_by_crypto(config):
    output_folder = config["output_folder"]
    output_report = config["output_report"]
    timeframes = config["timeframes"]
    start_year = config["start_year"]
    end_year = config["end_year"]

    os.makedirs(output_report, exist_ok=True)

    # ✅ SOMENTE arquivos *_strategy.xlsx
    files = glob.glob(os.path.join(output_folder, "*_strategy.xlsx"))

    for file in tqdm(files, desc="📊 Detailed Report", unit="crypto"):
        crypto = os.path.basename(file).replace("_strategy.xlsx", "")
        pdf_path = os.path.join(output_report, f"{crypto}_detailed.pdf")

        wb = load_workbook(file)

        with PdfPages(pdf_path) as pdf:
            for year in range(start_year, end_year + 1):
                for tf in timeframes:
                    if tf not in wb.sheetnames:
                        continue

                    df = pd.read_excel(file, sheet_name=tf)

                    if "Start" not in df.columns:
                        print(f"⚠️ Sheet {tf} in {file} has no 'Start' column")
                        continue

                    df["Date"] = pd.to_datetime(df["Start"])
                    df["Year"] = df["Date"].dt.year
                    df["Month"] = df["Date"].dt.month.map(MONTHS_EN)

                    df_year = df[df["Year"] == year]
                    if df_year.empty:
                        continue

                    selected_columns = [
                        "Month",
                        "Total Return [%]",
                        "Benchmark Return [%]",
                        "Total Trades",
                        "Total Closed Trades",
                        "Total Open Trades",
                        "Open Trade PnL"
                    ]

                    table = df_year[selected_columns].copy()
                    table["Month"] = pd.Categorical(
                        table["Month"],
                        categories=list(MONTHS_EN.values()),
                        ordered=True
                    )
                    table.sort_values(by="Month", inplace=True)

                    table["Total Return [%]"] = table["Total Return [%]"].map(lambda x: f"{x:.2f}%")
                    table["Benchmark Return [%]"] = table["Benchmark Return [%]"].map(lambda x: f"{x:.2f}%")
                    table["Open Trade PnL"] = table["Open Trade PnL"].map(lambda x: f"{x:.2f}")
                    table["Total Trades"] = table["Total Trades"].astype(int)
                    table["Total Closed Trades"] = table["Total Closed Trades"].astype(int)
                    table["Total Open Trades"] = table["Total Open Trades"].astype(int)

                    sum_row = {
                        "Month": "Total",
                        "Total Return [%]": f"{df_year['Total Return [%]'].sum():.2f}%",
                        "Benchmark Return [%]": f"{df_year['Benchmark Return [%]'].sum():.2f}%",
                        "Total Trades": df_year["Total Trades"].sum(),
                        "Total Closed Trades": df_year["Total Closed Trades"].sum(),
                        "Total Open Trades": df_year["Total Open Trades"].sum(),
                        "Open Trade PnL": f"{df_year['Open Trade PnL'].sum():.2f}"
                    }

                    table = pd.concat([table, pd.DataFrame([sum_row])], ignore_index=True)

                    fig, ax = plt.subplots(figsize=(10, 6))
                    ax.axis("off")
                    ax.axis("tight")

                    table_data = [table.columns.tolist()] + table.values.tolist()
                    table_plot = ax.table(
                        cellText=table_data,
                        colLabels=None,
                        loc="center",
                        cellLoc="center"
                    )
                    table_plot.scale(1.1, 1.3)

                    header_color = "#D3D3D3"
                    last_row = len(table_data) - 1
                    num_cols = len(table.columns)

                    for col in range(num_cols):
                        table_plot[0, col].set_facecolor(header_color)
                        table_plot[0, col].set_text_props(weight="bold")
                        table_plot[last_row, col].set_facecolor(header_color)
                        table_plot[last_row, col].set_text_props(weight="bold")

                    plt.title(f"{crypto} - {tf} - {year}", fontsize=14)
                    plt.tight_layout()
                    pdf.savefig()
                    plt.close()
