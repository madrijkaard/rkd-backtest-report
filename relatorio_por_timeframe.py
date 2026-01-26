import os
import glob
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import load_workbook
from tqdm import tqdm

MONTHS_EN = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
}

def generate_timeframe_report_by_crypto(config):
    output_folder = config["output_folder"]
    output_report = config["output_report"]
    timeframes = config["timeframes"]
    start_year = config["start_year"]
    end_year = config["end_year"]

    os.makedirs(output_report, exist_ok=True)
    files = glob.glob(os.path.join(output_folder, "*.xlsx"))

    for file in tqdm(files, desc="📈 Timeframe Report", unit="crypto"):
        crypto = os.path.basename(file).replace(".xlsx", "")
        wb = load_workbook(file)

        pdf_path = os.path.join(output_report, f"{crypto}_timeframe.pdf")
        with PdfPages(pdf_path) as pdf:
            for tf in timeframes:
                if tf not in wb.sheetnames:
                    continue

                df = pd.read_excel(file, sheet_name=tf)

                # ✅ Usar a coluna 'Start' como referência de data
                if "Start" not in df.columns:
                    print(f"⚠️ Sheet {tf} in {file} has no 'Start' column")
                    continue

                df["Date"] = pd.to_datetime(df["Start"])
                df["Year"] = df["Date"].dt.year
                df["Month"] = df["Date"].dt.month.map(MONTHS_EN)

                for year in range(start_year, end_year + 1):
                    df_year = df[df["Year"] == year]
                    if df_year.empty:
                        continue

                    monthly_return = df_year.groupby("Month")["Total Return [%]"].sum()
                    monthly_return = monthly_return.reindex(MONTHS_EN.values())

                    plt.figure(figsize=(10, 5))
                    monthly_return.plot(kind="bar", color="mediumseagreen", edgecolor="black")
                    plt.title(f"{crypto} - {tf} - Monthly Return - {year}")
                    plt.ylabel("Return [%]")
                    plt.xlabel("Month")
                    plt.axhline(0, color='black', linewidth=0.8)
                    plt.xticks(rotation=45)
                    plt.tight_layout()
                    pdf.savefig()
                    plt.close()
